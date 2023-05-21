# import python libraries
# parse the file
# select the root
# create an excel file

# set up a dictionary contain the parent child relationships

# call a function to return the total number of the child nodes

# set up 4 lists to store the data we needed
# find the terms with defstr containing 'autophagosome'
# Store the corresponding data in the lists

# write the data into the excel file created


# import necessary python libraries
import xml.dom.minidom
from openpyxl import workbook

# read the file and select 'term' data
DOMTree = xml.dom.minidom.parse('go_obo.xml')
root = DOMTree.documentElement
GO = root.getElementsByTagName('term')

# create an excel file
wb = workbook.Workbook()
sheet = wb.worksheets[0] # move to the first sheet

# set up the parent child relationship
# create a dictionary
Dic_relationship = {}
# find the id and the corresponding is_a information
for go in GO:
    id = go.getElementsByTagName('id')[0].firstChild.data
    Dic_relationship[id] =[]

for go in GO:
    id = go.getElementsByTagName('id')[0].firstChild.data
    is_a_list = go.getElementsByTagName('is_a')
    for is_a in is_a_list:
        p_id = is_a.firstChild.data
        Dic_relationship[p_id].append(id)

# call a function
def total_num_cn(id):
    count = 0 # 'count' to store the number of child nodes
    num_cn = len(Dic_relationship[id])
    count += num_cn
    if num_cn != 0: # if there is any child nodes
        cns = [] # to store the child nodes to use later to find child nodes of the child nodes
        for ids in Dic_relationship[id]:
            cns.extend(Dic_relationship[ids]) # store all child nodes from all the child nodes found
        num_cn = len(cns)
        count += num_cn # add the number of child nodes
        while num_cn != 0: # repeat until there is no any child nodes
            temp_list = []
            for cn in cns:
                temp_list.extend(Dic_relationship[cn])
            num_cn = len(temp_list)
            count += num_cn
            cns = temp_list
        return count # return the total number of child nodes
    else: # there is no child nodes return 0
        return 0


# create 4 lists to store the ids, names, defstr and number of child nodes
id_list = []
name_list = []
defstr_list = []
cn_list = []

# find and select the terms with 'defstr' containing 'autophagosome'
for go in GO:
    defstr = go.getElementsByTagName('defstr')[0].firstChild.data # extract the string in 'defstr'
    match = defstr.find('autophagosome')
    count = 0
    if match >= 0: # if there is autophagosome line 75 will return a positive number if not -1
        id = go.getElementsByTagName('id')[0].firstChild.data
        name = go.getElementsByTagName('name')[0].firstChild.data
        cn = total_num_cn(id)

        id_list.append(id)
        name_list.append(name)
        defstr_list.append(defstr)
        cn_list.append(cn)


# write the data obtained into the excel file
for i in range(len(id_list)):
    cell = sheet.cell(i+1,1)
    cell.value = id_list[i]
for i in range(len(name_list)):
    cell = sheet.cell(i+1,2)
    cell.value = name_list[i]
for i in range(len(defstr_list)):
    cell = sheet.cell(i+1,3)
    cell.value = defstr_list[i]
for i in range(len(cn_list)):
    cell = sheet.cell(i+1,4)
    cell.value = cn_list[i]
wb.save('autophagosome.xlsx')


